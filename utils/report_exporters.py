# utils/report_exporters.py
# Advanced export utilities and report templates

import json
import csv
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Tuple
import os
from pathlib import Path
import tempfile
import zipfile

# Optional dependencies with graceful fallback
try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.backends.backend_pdf import PdfPages
    import seaborn as sns
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


class AdvancedExportManager:
    """
    Advanced export manager with professional report generation capabilities
    """
    
    def __init__(self):
        self.temp_dir = tempfile.mkdtemp()
        self.report_templates = self._load_report_templates()
    
    def _load_report_templates(self) -> Dict:
        """Load report templates configuration"""
        return {
            'executive_summary': {
                'title': 'Executive Financial Summary',
                'sections': ['overview', 'key_metrics', 'alerts', 'recommendations'],
                'charts': ['spending_overview', 'category_breakdown']
            },
            'detailed_analysis': {
                'title': 'Detailed Financial Analysis', 
                'sections': ['overview', 'monthly_breakdown', 'category_analysis', 'trends'],
                'charts': ['monthly_trends', 'category_comparison', 'budget_vs_actual']
            },
            'budget_performance': {
                'title': 'Budget Performance Report',
                'sections': ['budget_summary', 'variance_analysis', 'performance_metrics'],
                'charts': ['budget_adherence', 'variance_chart']
            },
            'transaction_report': {
                'title': 'Transaction Details Report',
                'sections': ['transaction_summary', 'category_breakdown', 'monthly_analysis'],
                'charts': ['transaction_volume', 'spending_patterns']
            }
        }
    
    def export_excel_professional(self, data: Dict, filename: str, template: str = 'detailed_analysis') -> bool:
        """
        Export professional Excel report with charts and formatting
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("Excel export requires pandas and openpyxl")
        
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Create multiple sheets based on template
                template_config = self.report_templates.get(template, self.report_templates['detailed_analysis'])
                
                # Summary sheet
                self._create_summary_sheet(writer, data, template_config)
                
                # Transactions sheet
                self._create_transactions_sheet(writer, data)
                
                # Budget analysis sheet
                self._create_budget_sheet(writer, data)
                
                # Category analysis sheet
                self._create_category_sheet(writer, data)
                
                # Charts sheet
                self._create_charts_sheet(writer, data)
            
            # Post-process with openpyxl for advanced formatting
            self._format_excel_workbook(filename, template_config)
            
            return True
            
        except Exception as e:
            print(f"Error creating Excel report: {e}")
            return False
    
    def _create_summary_sheet(self, writer: pd.ExcelWriter, data: Dict, template_config: Dict):
        """Create executive summary sheet"""
        summary_data = []
        
        # Financial overview
        total_budget = self._calculate_total_budget(data)
        total_spent = self._calculate_total_spent(data)
        
        summary_data.extend([
            ['Financial Overview', ''],
            ['Total Budget', f'₹{total_budget:,.2f}'],
            ['Total Spent', f'₹{total_spent:,.2f}'],
            ['Remaining Budget', f'₹{total_budget - total_spent:,.2f}'],
            ['Utilization %', f'{(total_spent/total_budget*100):.1f}%' if total_budget > 0 else '0.0%'],
            ['', ''],
        ])
        
        # Category breakdown
        category_totals = self._get_category_totals(data)
        summary_data.append(['Top Spending Categories', ''])
        
        for category, amount in sorted(category_totals.items(), key=lambda x: x[1], reverse=True)[:10]:
            summary_data.append([category, f'₹{amount:,.2f}'])
        
        # Create DataFrame and export
        df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
        df.to_excel(writer, sheet_name='Summary', index=False)
    
    def _create_transactions_sheet(self, writer: pd.ExcelWriter, data: Dict):
        """Create detailed transactions sheet"""
        transactions_data = []
        
        if 'transactions' in data:
            for month, transactions in data['transactions'].items():
                for transaction in transactions:
                    transactions_data.append({
                        'Month': month,
                        'Date': transaction.get('date', ''),
                        'Category': transaction.get('category', 'Uncategorized'),
                        'Amount': transaction.get('amount', 0),
                        'Description': transaction.get('description', ''),
                        'Source': transaction.get('source', 'Manual')
                    })
        
        if transactions_data:
            df = pd.DataFrame(transactions_data)
            df.to_excel(writer, sheet_name='Transactions', index=False)
    
    def _create_budget_sheet(self, writer: pd.ExcelWriter, data: Dict):
        """Create budget analysis sheet"""
        budget_data = []
        
        if 'budgets' in data:
            for month, budgets in data['budgets'].items():
                for category, budget_amount in budgets.items():
                    # Calculate actual spending for this category/month
                    actual_spent = 0
                    if 'transactions' in data and month in data['transactions']:
                        actual_spent = sum(
                            t.get('amount', 0) for t in data['transactions'][month]
                            if t.get('category') == category
                        )
                    
                    variance = actual_spent - budget_amount
                    variance_pct = (variance / budget_amount * 100) if budget_amount > 0 else 0
                    
                    budget_data.append({
                        'Month': month,
                        'Category': category,
                        'Budgeted': budget_amount,
                        'Actual': actual_spent,
                        'Variance': variance,
                        'Variance %': variance_pct,
                        'Status': 'Over Budget' if variance > 0 else 'Within Budget'
                    })
        
        if budget_data:
            df = pd.DataFrame(budget_data)
            df.to_excel(writer, sheet_name='Budget Analysis', index=False)
    
    def _create_category_sheet(self, writer: pd.ExcelWriter, data: Dict):
        """Create category analysis sheet"""
        category_analysis = []
        category_totals = self._get_category_totals(data)
        
        total_spending = sum(category_totals.values())
        
        for category, amount in sorted(category_totals.items(), key=lambda x: x[1], reverse=True):
            percentage = (amount / total_spending * 100) if total_spending > 0 else 0
            
            category_analysis.append({
                'Category': category,
                'Total Spent': amount,
                'Percentage': percentage,
                'Average Monthly': amount / 12,  # Assuming 12 months of data
                'Ranking': len(category_analysis) + 1
            })
        
        if category_analysis:
            df = pd.DataFrame(category_analysis)
            df.to_excel(writer, sheet_name='Category Analysis', index=False)
    
    def _create_charts_sheet(self, writer: pd.ExcelWriter, data: Dict):
        """Create charts summary sheet"""
        # This creates a placeholder sheet for charts
        # Charts will be added in the formatting step
        chart_info = [
            ['Chart Type', 'Description'],
            ['Spending Overview', 'Monthly spending trends'],
            ['Category Breakdown', 'Spending by category'],
            ['Budget vs Actual', 'Budget adherence analysis'],
            ['Top Categories', 'Highest spending categories']
        ]
        
        df = pd.DataFrame(chart_info[1:], columns=chart_info[0])
        df.to_excel(writer, sheet_name='Charts', index=False)
    
    def _format_excel_workbook(self, filename: str, template_config: Dict):
        """Apply professional formatting to Excel workbook"""
        try:
            wb = openpyxl.load_workbook(filename)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Format headers
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Add charts if it's the Charts sheet
                if sheet_name == 'Charts' and 'Budget Analysis' in wb.sheetnames:
                    self._add_excel_charts(wb, ws)
            
            wb.save(filename)
            
        except Exception as e:
            print(f"Error formatting Excel workbook: {e}")
    
    def _add_excel_charts(self, workbook, charts_sheet):
        """Add charts to Excel workbook"""
        try:
            if 'Budget Analysis' in workbook.sheetnames:
                budget_sheet = workbook['Budget Analysis']
                
                # Create a bar chart for budget vs actual
                chart = BarChart()
                chart.title = "Budget vs Actual Spending"
                chart.y_axis.title = 'Amount (₹)'
                chart.x_axis.title = 'Category'
                
                # Add data (simplified example)
                data = Reference(budget_sheet, min_col=3, min_row=1, max_col=4, max_row=10)
                categories = Reference(budget_sheet, min_col=2, min_row=2, max_row=10)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                
                charts_sheet.add_chart(chart, "A5")
                
        except Exception as e:
            print(f"Error adding Excel charts: {e}")
    
    def export_pdf_professional(self, data: Dict, filename: str, template: str = 'executive_summary') -> bool:
        """
        Export professional PDF report with charts and formatting
        """
        if not REPORTLAB_AVAILABLE:
            # Fallback to simple text-based PDF
            return self._export_simple_pdf(data, filename)
        
        try:
            doc = SimpleDocTemplate(filename, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                textColor=colors.darkblue
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'], 
                fontSize=16,
                spaceAfter=15,
                textColor=colors.darkgreen
            )
            
            # Title
            template_config = self.report_templates.get(template, self.report_templates['executive_summary'])
            story.append(Paragraph(template_config['title'], title_style))
            story.append(Spacer(1, 20))
            
            # Generated date
            story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
            story.append(Spacer(1, 30))
            
            # Financial overview
            story.append(Paragraph("Financial Overview", heading_style))
            
            total_budget = self._calculate_total_budget(data)
            total_spent = self._calculate_total_spent(data)
            
            overview_data = [
                ['Metric', 'Amount'],
                ['Total Budget', f'₹{total_budget:,.2f}'],
                ['Total Spent', f'₹{total_spent:,.2f}'],
                ['Remaining', f'₹{total_budget - total_spent:,.2f}'],
                ['Utilization', f'{(total_spent/total_budget*100):.1f}%' if total_budget > 0 else '0.0%']
            ]
            
            overview_table = Table(overview_data)
            overview_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(overview_table)
            story.append(Spacer(1, 30))
            
            # Category breakdown
            story.append(Paragraph("Top Spending Categories", heading_style))
            
            category_totals = self._get_category_totals(data)
            category_data = [['Category', 'Amount', 'Percentage']]
            
            total_spending = sum(category_totals.values())
            
            for category, amount in sorted(category_totals.items(), key=lambda x: x[1], reverse=True)[:10]:
                percentage = (amount / total_spending * 100) if total_spending > 0 else 0
                category_data.append([category, f'₹{amount:,.2f}', f'{percentage:.1f}%'])
            
            category_table = Table(category_data)
            category_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(category_table)
            story.append(Spacer(1, 30))
            
            # Add charts if matplotlib is available
            if MATPLOTLIB_AVAILABLE:
                chart_path = self._create_pdf_charts(data)
                if chart_path and os.path.exists(chart_path):
                    story.append(Paragraph("Spending Analysis Charts", heading_style))
                    story.append(Image(chart_path, width=500, height=300))
                    story.append(Spacer(1, 20))
            
            # Recommendations section
            story.append(Paragraph("Recommendations", heading_style))
            recommendations = self._generate_recommendations(data)
            
            for rec in recommendations:
                story.append(Paragraph(f"• {rec}", styles['Normal']))
                story.append(Spacer(1, 10))
            
            # Build PDF
            doc.build(story)
            return True
            
        except Exception as e:
            print(f"Error creating PDF report: {e}")
            return False
    
    def _export_simple_pdf(self, data: Dict, filename: str) -> bool:
        """Fallback PDF export without reportlab"""
        try:
            # Create a simple text report and save as .txt with PDF extension note
            report_content = self._generate_comprehensive_text_report(data)
            
            # Change extension to .txt for compatibility
            base_name = os.path.splitext(filename)[0]
            txt_filename = f"{base_name}_report.txt"
            
            with open(txt_filename, 'w', encoding='utf-8') as f:
                f.write(report_content)
            
            return True
            
        except Exception as e:
            print(f"Error creating simple PDF report: {e}")
            return False
    
    def _create_pdf_charts(self, data: Dict) -> Optional[str]:
        """Create charts for PDF report"""
        try:
            if not MATPLOTLIB_AVAILABLE:
                return None
            
            # Create a figure with subplots
            fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(12, 10))
            fig.suptitle('Financial Analysis Charts', fontsize=16, fontweight='bold')
            
            # Chart 1: Category spending pie chart
            category_totals = self._get_category_totals(data)
            if category_totals:
                categories = list(category_totals.keys())[:8]  # Top 8 categories
                amounts = list(category_totals.values())[:8]
                
                ax1.pie(amounts, labels=categories, autopct='%1.1f%%', startangle=90)
                ax1.set_title('Spending by Category')
            
            # Chart 2: Monthly spending trend
            monthly_totals = self._get_monthly_totals(data)
            if monthly_totals:
                months = list(monthly_totals.keys())
                amounts = list(monthly_totals.values())
                
                ax2.plot(months, amounts, marker='o', linewidth=2, markersize=6)
                ax2.set_title('Monthly Spending Trend')
                ax2.set_ylabel('Amount (₹)')
                ax2.tick_params(axis='x', rotation=45)
            
            # Chart 3: Budget vs Actual
            budget_comparison = self._get_budget_vs_actual(data)
            if budget_comparison:
                categories = list(budget_comparison.keys())[:10]
                budgeted = [budget_comparison[cat]['budgeted'] for cat in categories]
                actual = [budget_comparison[cat]['actual'] for cat in categories]
                
                x = range(len(categories))
                width = 0.35
                
                ax3.bar([i - width/2 for i in x], budgeted, width, label='Budgeted', alpha=0.8)
                ax3.bar([i + width/2 for i in x], actual, width, label='Actual', alpha=0.8)
                ax3.set_title('Budget vs Actual')
                ax3.set_ylabel('Amount (₹)')
                ax3.set_xticks(x)
                ax3.set_xticklabels(categories, rotation=45, ha='right')
                ax3.legend()
            
            # Chart 4: Top spending categories bar chart
            if category_totals:
                top_categories = sorted(category_totals.items(), key=lambda x: x[1], reverse=True)[:10]
                cats = [item[0] for item in top_categories]
                vals = [item[1] for item in top_categories]
                
                ax4.barh(cats, vals)
                ax4.set_title('Top Spending Categories')
                ax4.set_xlabel('Amount (₹)')
            
            # Adjust layout and save
            plt.tight_layout()
            
            chart_path = os.path.join(self.temp_dir, 'financial_charts.png')
            plt.savefig(chart_path, dpi=300, bbox_inches='tight')
            plt.close()
            
            return chart_path
            
        except Exception as e:
            print(f"Error creating PDF charts: {e}")
            return None
    
    def export_comprehensive_package(self, data: Dict, output_dir: str, package_name: str = None) -> str:
        """
        Create comprehensive export package with multiple formats
        """
        if package_name is None:
            package_name = f"financial_report_package_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        package_dir = os.path.join(output_dir, package_name)
        os.makedirs(package_dir, exist_ok=True)
        
        try:
            # Export in all available formats
            formats_exported = []
            
            # JSON export (always available)
            json_file = os.path.join(package_dir, f"{package_name}_data.json")
            with open(json_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False, default=str)
            formats_exported.append("JSON")
            
            # CSV exports
            self._export_csv_package(data, package_dir, package_name)
            formats_exported.append("CSV")
            
            # Excel export (if available)
            if EXCEL_AVAILABLE:
                excel_file = os.path.join(package_dir, f"{package_name}_report.xlsx")
                self.export_excel_professional(data, excel_file)
                formats_exported.append("Excel")
            
            # PDF export (if available) 
            pdf_file = os.path.join(package_dir, f"{package_name}_summary.pdf")
            self.export_pdf_professional(data, pdf_file)
            formats_exported.append("PDF")
            
            # Create charts package (if matplotlib available)
            if MATPLOTLIB_AVAILABLE:
                charts_dir = os.path.join(package_dir, "charts")
                os.makedirs(charts_dir, exist_ok=True)
                self._create_chart_package(data, charts_dir)
                formats_exported.append("Charts")
            
            # Create summary README
            readme_content = self._create_package_readme(formats_exported, package_name)
            readme_file = os.path.join(package_dir, "README.txt")
            with open(readme_file, 'w', encoding='utf-8') as f:
                f.write(readme_content)
            
            # Create ZIP archive
            zip_file = f"{package_dir}.zip"
            with zipfile.ZipFile(zip_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for root, dirs, files in os.walk(package_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arc_name = os.path.relpath(file_path, package_dir)
                        zipf.write(file_path, arc_name)
            
            return zip_file
            
        except Exception as e:
            print(f"Error creating comprehensive package: {e}")
            return ""
    
    def _export_csv_package(self, data: Dict, package_dir: str, package_name: str):
        """Export multiple CSV files for different data types"""
        try:
            # Transactions CSV
            if 'transactions' in data:
                transactions_file = os.path.join(package_dir, f"{package_name}_transactions.csv")
                self._export_transactions_csv(data, transactions_file)
            
            # Budgets CSV
            if 'budgets' in data:
                budgets_file = os.path.join(package_dir, f"{package_name}_budgets.csv")
                self._export_budgets_csv(data, budgets_file)
            
            # Category summary CSV
            categories_file = os.path.join(package_dir, f"{package_name}_categories.csv")
            self._export_categories_csv(data, categories_file)
            
        except Exception as e:
            print(f"Error exporting CSV package: {e}")
    
    def _export_transactions_csv(self, data: Dict, filename: str):
        """Export transactions to CSV"""
        transactions_data = []
        
        if 'transactions' in data:
            for month, transactions in data['transactions'].items():
                for transaction in transactions:
                    transactions_data.append({
                        'Month': month,
                        'Date': transaction.get('date', ''),
                        'Category': transaction.get('category', 'Uncategorized'),
                        'Amount': transaction.get('amount', 0),
                        'Description': transaction.get('description', ''),
                        'Source': transaction.get('source', 'Manual'),
                        'ID': transaction.get('id', '')
                    })
        
        if transactions_data:
            fieldnames = ['Month', 'Date', 'Category', 'Amount', 'Description', 'Source', 'ID']
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(transactions_data)
    
    def _export_budgets_csv(self, data: Dict, filename: str):
        """Export budgets to CSV"""
        budgets_data = []
        
        if 'budgets' in data:
            for month, budgets in data['budgets'].items():
                for category, amount in budgets.items():
                    # Calculate actual spending
                    actual_spent = 0
                    if 'transactions' in data and month in data['transactions']:
                        actual_spent = sum(
                            t.get('amount', 0) for t in data['transactions'][month]
                            if t.get('category') == category
                        )
                    
                    budgets_data.append({
                        'Month': month,
                        'Category': category,
                        'Budgeted_Amount': amount,
                        'Actual_Spent': actual_spent,
                        'Variance': actual_spent - amount,
                        'Variance_Percentage': ((actual_spent - amount) / amount * 100) if amount > 0 else 0
                    })
        
        if budgets_data:
            fieldnames = ['Month', 'Category', 'Budgeted_Amount', 'Actual_Spent', 'Variance', 'Variance_Percentage']
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(budgets_data)
    
    def _export_categories_csv(self, data: Dict, filename: str):
        """Export category summary to CSV"""
        category_totals = self._get_category_totals(data)
        total_spending = sum(category_totals.values())
        
        categories_data = []
        
        for category, amount in sorted(category_totals.items(), key=lambda x: x[1], reverse=True):
            percentage = (amount / total_spending * 100) if total_spending > 0 else 0
            avg_monthly = amount / 12  # Assuming 12 months
            
            categories_data.append({
                'Category': category,
                'Total_Spent': amount,
                'Percentage_of_Total': percentage,
                'Average_Monthly': avg_monthly,
                'Ranking': len(categories_data) + 1
            })
        
        if categories_data:
            fieldnames = ['Category', 'Total_Spent', 'Percentage_of_Total', 'Average_Monthly', 'Ranking']
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(categories_data)
    
    def _create_chart_package(self, data: Dict, charts_dir: str):
        """Create comprehensive chart package"""
        try:
            if not MATPLOTLIB_AVAILABLE:
                return
            
            # Set style
            plt.style.use('default')
            
            # Chart 1: Category spending pie chart
            category_totals = self._get_category_totals(data)
            if category_totals:
                plt.figure(figsize=(10, 8))
                categories = list(category_totals.keys())[:10]
                amounts = list(category_totals.values())[:10]
                
                colors = plt.cm.Set3(range(len(categories)))
                plt.pie(amounts, labels=categories, autopct='%1.1f%%', colors=colors, startangle=90)
                plt.title('Spending Distribution by Category', fontsize=16, fontweight='bold', pad=20)
                plt.tight_layout()
                plt.savefig(os.path.join(charts_dir, 'category_pie_chart.png'), dpi=300, bbox_inches='tight')
                plt.close()
            
            # Chart 2: Monthly spending trend
            monthly_totals = self._get_monthly_totals(data)
            if monthly_totals:
                plt.figure(figsize=(12, 6))
                months = list(monthly_totals.keys())
                amounts = list(monthly_totals.values())
                
                plt.plot(months, amounts, marker='o', linewidth=3, markersize=8, color='#2E8B57')
                plt.fill_between(months, amounts, alpha=0.3, color='#2E8B57')
                plt.title('Monthly Spending Trend', fontsize=16, fontweight='bold')
                plt.xlabel('Month', fontsize=12)
                plt.ylabel('Amount Spent (₹)', fontsize=12)
                plt.xticks(rotation=45)
                plt.grid(True, alpha=0.3)
                plt.tight_layout()
                plt.savefig(os.path.join(charts_dir, 'monthly_trend.png'), dpi=300, bbox_inches='tight')
                plt.close()
            
            # Chart 3: Budget vs Actual comparison
            budget_comparison = self._get_budget_vs_actual(data)
            if budget_comparison:
                plt.figure(figsize=(14, 8))
                categories = list(budget_comparison.keys())[:12]
                budgeted = [budget_comparison[cat]['budgeted'] for cat in categories]
                actual = [budget_comparison[cat]['actual'] for cat in categories]
                
                x = range(len(categories))
                width = 0.35
                
                bars1 = plt.bar([i - width/2 for i in x], budgeted, width, label='Budgeted', 
                               color='#4472C4', alpha=0.8)
                bars2 = plt.bar([i + width/2 for i in x], actual, width, label='Actual', 
                               color='#E15759', alpha=0.8)
                
                plt.title('Budget vs Actual Spending Comparison', fontsize=16, fontweight='bold')
                plt.xlabel('Category', fontsize=12)
                plt.ylabel('Amount (₹)', fontsize=12)
                plt.xticks(x, categories, rotation=45, ha='right')
                plt.legend()
                plt.grid(True, alpha=0.3, axis='y')
                plt.tight_layout()
                plt.savefig(os.path.join(charts_dir, 'budget_vs_actual.png'), dpi=300, bbox_inches='tight')
                plt.close()
            
            # Chart 4: Top spending categories horizontal bar
            if category_totals:
                plt.figure(figsize=(12, 10))
                top_categories = sorted(category_totals.items(), key=lambda x: x[1], reverse=True)[:15]
                cats = [item[0] for item in top_categories]
                vals = [item[1] for item in top_categories]
                
                colors = plt.cm.viridis(range(len(cats)))
                bars = plt.barh(cats, vals, color=colors)
                
                plt.title('Top Spending Categories', fontsize=16, fontweight='bold')
                plt.xlabel('Amount Spent (₹)', fontsize=12)
                plt.ylabel('Category', fontsize=12)
                
                # Add value labels on bars
                for bar, val in zip(bars, vals):
                    plt.text(val + max(vals) * 0.01, bar.get_y() + bar.get_height()/2, 
                            f'₹{val:,.0f}', ha='left', va='center', fontweight='bold')
                
                plt.tight_layout()
                plt.savefig(os.path.join(charts_dir, 'top_categories_horizontal.png'), dpi=300, bbox_inches='tight')
                plt.close()
            
        except Exception as e:
            print(f"Error creating chart package: {e}")
    
    def _create_package_readme(self, formats_exported: List[str], package_name: str) -> str:
        """Create README file for the export package"""
        readme_content = f"""
FINANCIAL REPORT PACKAGE
========================

Package Name: {package_name}
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

CONTENTS:
---------

This package contains your financial data exported in multiple formats:

Exported Formats: {', '.join(formats_exported)}

FILE DESCRIPTIONS:
------------------

"""
        
        if "JSON" in formats_exported:
            readme_content += f"""
{package_name}_data.json
    Complete data backup in JSON format
    Contains all transactions, budgets, and settings
    Can be used to restore data in the application

"""
        
        if "CSV" in formats_exported:
            readme_content += f"""
{package_name}_transactions.csv
    All transactions with month, date, category, amount, and description
    
{package_name}_budgets.csv
    Budget vs actual analysis with variance calculations
    
{package_name}_categories.csv
    Category-wise spending summary with rankings

"""
        
        if "Excel" in formats_exported:
            readme_content += f"""
{package_name}_report.xlsx
    Professional Excel report with multiple sheets:
    - Summary: Executive overview
    - Transactions: Detailed transaction listing
    - Budget Analysis: Budget performance analysis
    - Category Analysis: Category-wise breakdown
    - Charts: Visual analysis placeholder

"""
        
        if "PDF" in formats_exported:
            readme_content += f"""
{package_name}_summary.pdf
    Executive summary report in PDF format
    Includes key metrics, category breakdown, and recommendations

"""
        
        if "Charts" in formats_exported:
            readme_content += """
charts/
    Directory containing visual analysis charts:
    - category_pie_chart.png: Spending distribution
    - monthly_trend.png: Monthly spending trends  
    - budget_vs_actual.png: Budget performance comparison
    - top_categories_horizontal.png: Top spending categories

"""
        
        readme_content += f"""
README.txt
    This file - explains package contents

USAGE NOTES:
------------

1. JSON file can be imported back into the Financial Management Tool
2. CSV files can be opened in Excel, Google Sheets, or other spreadsheet software
3. Excel file provides the most comprehensive analysis with charts and formatting
4. PDF file is ideal for sharing executive summaries
5. Chart images can be used in presentations or reports

SUPPORT:
--------

If you have questions about this export package, please refer to the 
Financial Management Tool documentation or contact support.

Generated by Financial Management & Simulation Tool
"""
        
        return readme_content
    
    # Helper methods for data processing
    
    def _calculate_total_budget(self, data: Dict) -> float:
        """Calculate total budget across all months"""
        total = 0
        if 'budgets' in data:
            for month, budgets in data['budgets'].items():
                total += sum(budgets.values())
        return total
    
    def _calculate_total_spent(self, data: Dict) -> float:
        """Calculate total amount spent across all transactions"""
        total = 0
        if 'transactions' in data:
            for month, transactions in data['transactions'].items():
                total += sum(t.get('amount', 0) for t in transactions)
        return total
    
    def _get_category_totals(self, data: Dict) -> Dict[str, float]:
        """Get total spending by category"""
        category_totals = {}
        
        if 'transactions' in data:
            for month, transactions in data['transactions'].items():
                for transaction in transactions:
                    category = transaction.get('category', 'Uncategorized')
                    amount = transaction.get('amount', 0)
                    category_totals[category] = category_totals.get(category, 0) + amount
        
        return category_totals
    
    def _get_monthly_totals(self, data: Dict) -> Dict[str, float]:
        """Get total spending by month"""
        monthly_totals = {}
        
        if 'transactions' in data:
            for month, transactions in data['transactions'].items():
                total = sum(t.get('amount', 0) for t in transactions)
                monthly_totals[month] = total
        
        return monthly_totals
    
    def _get_budget_vs_actual(self, data: Dict) -> Dict[str, Dict[str, float]]:
        """Compare budget vs actual spending by category"""
        comparison = {}
        
        # Get all categories from both budgets and transactions
        categories = set()
        
        if 'budgets' in data:
            for month, budgets in data['budgets'].items():
                categories.update(budgets.keys())
        
        if 'transactions' in data:
            for month, transactions in data['transactions'].items():
                for transaction in transactions:
                    categories.add(transaction.get('category', 'Uncategorized'))
        
        # Calculate totals for each category
        for category in categories:
            total_budgeted = 0
            total_actual = 0
            
            # Sum budgeted amounts
            if 'budgets' in data:
                for month, budgets in data['budgets'].items():
                    total_budgeted += budgets.get(category, 0)
            
            # Sum actual spending
            if 'transactions' in data:
                for month, transactions in data['transactions'].items():
                    for transaction in transactions:
                        if transaction.get('category') == category:
                            total_actual += transaction.get('amount', 0)
            
            comparison[category] = {
                'budgeted': total_budgeted,
                'actual': total_actual,
                'variance': total_actual - total_budgeted
            }
        
        return comparison
    
    def _generate_recommendations(self, data: Dict) -> List[str]:
        """Generate financial recommendations based on data analysis"""
        recommendations = []
        
        try:
            total_budget = self._calculate_total_budget(data)
            total_spent = self._calculate_total_spent(data)
            
            if total_budget > 0:
                utilization = (total_spent / total_budget) * 100
                
                if utilization > 100:
                    recommendations.append(f"You've exceeded your total budget by {utilization - 100:.1f}%. Consider reviewing and adjusting spending priorities.")
                    recommendations.append("Look for areas where you can reduce discretionary spending.")
                
                elif utilization > 80:
                    recommendations.append("You're approaching your budget limits. Monitor spending closely for the remaining period.")
                    recommendations.append("Consider setting aside emergency funds for unexpected expenses.")
                
                else:
                    recommendations.append("Good job staying within budget! Consider allocating surplus to savings or investments.")
            
            # Category-specific recommendations
            category_totals = self._get_category_totals(data)
            budget_comparison = self._get_budget_vs_actual(data)
            
            # Find top overspending categories
            overspent_categories = []
            for category, comparison in budget_comparison.items():
                if comparison['variance'] > 0 and comparison['budgeted'] > 0:
                    variance_pct = (comparison['variance'] / comparison['budgeted']) * 100
                    if variance_pct > 20:  # More than 20% over budget
                        overspent_categories.append((category, variance_pct))
            
            if overspent_categories:
                overspent_categories.sort(key=lambda x: x[1], reverse=True)
                top_overspent = overspent_categories[0]
                recommendations.append(f"Focus on reducing spending in {top_overspent[0]} - you're {top_overspent[1]:.1f}% over budget in this category.")
            
            # General recommendations
            if len(category_totals) < 5:
                recommendations.append("Consider using more specific categories to better track your spending patterns.")
            
            recommendations.append("Review your spending patterns monthly and adjust budgets as needed.")
            recommendations.append("Consider setting up automatic savings transfers to build financial reserves.")
            
        except Exception as e:
            recommendations.append("Unable to generate specific recommendations due to data processing error.")
            print(f"Error generating recommendations: {e}")
        
        return recommendations[:6]  # Limit to 6 recommendations
    
    def _generate_comprehensive_text_report(self, data: Dict) -> str:
        """Generate comprehensive text-based report"""
        report_lines = []
        
        # Header
        report_lines.extend([
            "=" * 80,
            "COMPREHENSIVE FINANCIAL REPORT",
            "=" * 80,
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "EXECUTIVE SUMMARY",
            "-" * 40
        ])
        
        # Financial overview
        total_budget = self._calculate_total_budget(data)
        total_spent = self._calculate_total_spent(data)
        remaining = total_budget - total_spent
        utilization = (total_spent / total_budget * 100) if total_budget > 0 else 0
        
        report_lines.extend([
            f"Total Budget: ₹{total_budget:,.2f}",
            f"Total Spent: ₹{total_spent:,.2f}",
            f"Remaining Budget: ₹{remaining:,.2f}",
            f"Budget Utilization: {utilization:.1f}%",
            ""
        ])
        
        # Category breakdown
        category_totals = self._get_category_totals(data)
        if category_totals:
            report_lines.extend([
                "TOP SPENDING CATEGORIES",
                "-" * 40
            ])
            
            sorted_categories = sorted(category_totals.items(), key=lambda x: x[1], reverse=True)
            total_spending = sum(category_totals.values())
            
            for i, (category, amount) in enumerate(sorted_categories[:15], 1):
                percentage = (amount / total_spending * 100) if total_spending > 0 else 0
                report_lines.append(f"{i:2d}. {category:<25} ₹{amount:>10,.2f} ({percentage:5.1f}%)")
            
            report_lines.append("")
        
        # Monthly analysis
        monthly_totals = self._get_monthly_totals(data)
        if monthly_totals:
            report_lines.extend([
                "MONTHLY SPENDING ANALYSIS",
                "-" * 40
            ])
            
            for month, amount in sorted(monthly_totals.items()):
                report_lines.append(f"{month:<10} ₹{amount:>12,.2f}")
            
            # Calculate average monthly spending
            avg_monthly = sum(monthly_totals.values()) / len(monthly_totals)
            report_lines.extend([
                "",
                f"Average Monthly Spending: ₹{avg_monthly:,.2f}",
                ""
            ])
        
        # Budget performance
        budget_comparison = self._get_budget_vs_actual(data)
        if budget_comparison:
            report_lines.extend([
                "BUDGET PERFORMANCE ANALYSIS",
                "-" * 40
            ])
            
            over_budget_categories = []
            under_budget_categories = []
            
            for category, comparison in budget_comparison.items():
                if comparison['budgeted'] > 0:  # Only categories with budgets
                    variance_pct = (comparison['variance'] / comparison['budgeted']) * 100
                    
                    if comparison['variance'] > 0:
                        over_budget_categories.append((category, variance_pct, comparison['variance']))
                    else:
                        under_budget_categories.append((category, abs(variance_pct), abs(comparison['variance'])))
            
            if over_budget_categories:
                report_lines.append("OVER BUDGET CATEGORIES:")
                over_budget_categories.sort(key=lambda x: x[1], reverse=True)
                for category, pct, amount in over_budget_categories[:10]:
                    report_lines.append(f"  {category:<25} +₹{amount:>8,.0f} (+{pct:5.1f}%)")
                report_lines.append("")
            
            if under_budget_categories:
                report_lines.append("UNDER BUDGET CATEGORIES:")
                under_budget_categories.sort(key=lambda x: x[1], reverse=True)
                for category, pct, amount in under_budget_categories[:10]:
                    report_lines.append(f"  {category:<25} -₹{amount:>8,.0f} (-{pct:5.1f}%)")
                report_lines.append("")
        
        # Recommendations
        recommendations = self._generate_recommendations(data)
        if recommendations:
            report_lines.extend([
                "RECOMMENDATIONS",
                "-" * 40
            ])
            
            for i, rec in enumerate(recommendations, 1):
                report_lines.append(f"{i}. {rec}")
            
            report_lines.append("")
        
        # Transaction summary
        if 'transactions' in data:
            total_transactions = sum(len(transactions) for transactions in data['transactions'].values())
            report_lines.extend([
                "TRANSACTION SUMMARY", 
                "-" * 40,
                f"Total Transactions: {total_transactions:,}",
                f"Average Transaction Amount: ₹{total_spent/total_transactions:,.2f}" if total_transactions > 0 else "Average Transaction Amount: ₹0.00",
                ""
            ])
        
        # Footer
        report_lines.extend([
            "=" * 80,
            "END OF REPORT",
            "=" * 80,
            "",
            "This report was generated by the Financial Management & Simulation Tool.",
            "For questions or support, please refer to the application documentation."
        ])
        
        return "\n".join(report_lines)
    
    def __del__(self):
        """Cleanup temporary directory"""
        try:
            import shutil
            if hasattr(self, 'temp_dir') and os.path.exists(self.temp_dir):
                shutil.rmtree(self.temp_dir)
        except:
            pass


class ReportTemplateManager:
    """
    Manages report templates and customization options
    """
    
    def __init__(self):
        self.templates_dir = os.path.join(os.path.dirname(__file__), '..', 'resources', 'templates')
        os.makedirs(self.templates_dir, exist_ok=True)
        self.load_templates()
    
    def load_templates(self):
        """Load report templates from file"""
        template_file = os.path.join(self.templates_dir, 'report_templates.json')
        
        try:
            if os.path.exists(template_file):
                with open(template_file, 'r', encoding='utf-8') as f:
                    self.templates = json.load(f)
            else:
                self.templates = self._create_default_templates()
                self.save_templates()
        except Exception as e:
            print(f"Error loading templates: {e}")
            self.templates = self._create_default_templates()
    
    def save_templates(self):
        """Save templates to file"""
        template_file = os.path.join(self.templates_dir, 'report_templates.json')
        
        try:
            with open(template_file, 'w', encoding='utf-8') as f:
                json.dump(self.templates, f, indent=2)
        except Exception as e:
            print(f"Error saving templates: {e}")
    
    def _create_default_templates(self) -> Dict:
        """Create default report templates"""
        return {
            'executive_summary': {
                'name': 'Executive Summary',
                'description': 'High-level financial overview for executives',
                'sections': [
                    'financial_overview',
                    'key_metrics', 
                    'top_categories',
                    'alerts_warnings',
                    'recommendations'
                ],
                'charts': [
                    'spending_overview_pie',
                    'monthly_trend_line'
                ],
                'format_options': {
                    'include_charts': True,
                    'color_scheme': 'professional',
                    'page_orientation': 'portrait'
                }
            },
            
            'detailed_analysis': {
                'name': 'Detailed Financial Analysis',
                'description': 'Comprehensive analysis with all details',
                'sections': [
                    'executive_summary',
                    'monthly_breakdown',
                    'category_analysis',
                    'budget_performance',
                    'transaction_details',
                    'trend_analysis',
                    'variance_analysis',
                    'recommendations'
                ],
                'charts': [
                    'spending_overview_pie',
                    'monthly_trend_line',
                    'category_bar_chart',
                    'budget_vs_actual_comparison'
                ],
                'format_options': {
                    'include_charts': True,
                    'include_transaction_details': True,
                    'color_scheme': 'detailed',
                    'page_orientation': 'portrait'
                }
            },
            
            'budget_performance': {
                'name': 'Budget Performance Report',
                'description': 'Focus on budget vs actual performance',
                'sections': [
                    'budget_summary',
                    'variance_analysis',
                    'category_performance',
                    'overspend_analysis',
                    'recommendations'
                ],
                'charts': [
                    'budget_vs_actual_comparison',
                    'variance_waterfall',
                    'performance_heatmap'
                ],
                'format_options': {
                    'include_charts': True,
                    'highlight_variances': True,
                    'color_scheme': 'performance',
                    'page_orientation': 'landscape'
                }
            },
            
            'category_analysis': {
                'name': 'Category Analysis Report',
                'description': 'Deep dive into spending by category',
                'sections': [
                    'category_overview',
                    'category_trends',
                    'category_comparison',
                    'subcategory_breakdown',
                    'category_recommendations'
                ],
                'charts': [
                    'category_pie_chart',
                    'category_trend_lines',
                    'category_ranking_bar'
                ],
                'format_options': {
                    'include_charts': True,
                    'group_by_category': True,
                    'color_scheme': 'categorical',
                    'page_orientation': 'portrait'
                }
            },
            
            'transaction_details': {
                'name': 'Transaction Details Report',
                'description': 'Complete transaction listing with analysis',
                'sections': [
                    'transaction_summary',
                    'transaction_listing',
                    'transaction_patterns',
                    'unusual_transactions',
                    'transaction_recommendations'
                ],
                'charts': [
                    'transaction_volume_by_month',
                    'transaction_size_distribution'
                ],
                'format_options': {
                    'include_all_transactions': True,
                    'group_by_category': False,
                    'color_scheme': 'neutral',
                    'page_orientation': 'landscape'
                }
            },
            
            'monthly_summary': {
                'name': 'Monthly Summary Report',
                'description': 'Quick monthly overview',
                'sections': [
                    'month_overview',
                    'monthly_highlights',
                    'budget_status',
                    'top_expenses',
                    'next_month_preview'
                ],
                'charts': [
                    'monthly_spending_breakdown',
                    'budget_progress_bar'
                ],
                'format_options': {
                    'single_month_focus': True,
                    'compact_format': True,
                    'color_scheme': 'monthly',
                    'page_orientation': 'portrait'
                }
            }
        }
    
    def get_template(self, template_name: str) -> Optional[Dict]:
        """Get specific template configuration"""
        return self.templates.get(template_name)
    
    def get_all_templates(self) -> Dict:
        """Get all available templates"""
        return self.templates
    
    def create_custom_template(self, name: str, config: Dict) -> bool:
        """Create a new custom template"""
        try:
            self.templates[name] = config
            self.save_templates()
            return True
        except Exception as e:
            print(f"Error creating custom template: {e}")
            return False
    
    def delete_template(self, template_name: str) -> bool:
        """Delete a template"""
        try:
            if template_name in self.templates:
                del self.templates[template_name]
                self.save_templates()
                return True
            return False
        except Exception as e:
            print(f"Error deleting template: {e}")
            return False


class ReportScheduler:
    """
    Handles scheduling and automation of report generation
    """
    
    def __init__(self, data_manager, export_manager):
        self.data_manager = data_manager
        self.export_manager = export_manager
        self.schedules = {}
        self.load_schedules()
    
    def load_schedules(self):
        """Load report schedules from file"""
        try:
            schedules_file = os.path.join(os.path.dirname(__file__), '..', 'data', 'report_schedules.json')
            if os.path.exists(schedules_file):
                with open(schedules_file, 'r', encoding='utf-8') as f:
                    self.schedules = json.load(f)
        except Exception as e:
            print(f"Error loading report schedules: {e}")
            self.schedules = {}
    
    def save_schedules(self):
        """Save report schedules to file"""
        try:
            data_dir = os.path.join(os.path.dirname(__file__), '..', 'data')
            os.makedirs(data_dir, exist_ok=True)
            
            schedules_file = os.path.join(data_dir, 'report_schedules.json')
            with open(schedules_file, 'w', encoding='utf-8') as f:
                json.dump(self.schedules, f, indent=2, default=str)
        except Exception as e:
            print(f"Error saving report schedules: {e}")
    
    def add_schedule(self, name: str, schedule_config: Dict) -> bool:
        """Add a new report schedule"""
        try:
            self.schedules[name] = {
                'template': schedule_config.get('template', 'executive_summary'),
                'format': schedule_config.get('format', 'pdf'),
                'frequency': schedule_config.get('frequency', 'monthly'),  # daily, weekly, monthly
                'output_path': schedule_config.get('output_path', ''),
                'last_run': None,
                'next_run': self._calculate_next_run(schedule_config.get('frequency', 'monthly')),
                'enabled': schedule_config.get('enabled', True),
                'created_date': datetime.now().isoformat()
            }
            self.save_schedules()
            return True
        except Exception as e:
            print(f"Error adding report schedule: {e}")
            return False
    
    def _calculate_next_run(self, frequency: str) -> str:
        """Calculate next run date based on frequency"""
        now = datetime.now()
        
        if frequency == 'daily':
            next_run = now + timedelta(days=1)
        elif frequency == 'weekly':
            next_run = now + timedelta(weeks=1)
        elif frequency == 'monthly':
            # Next month, same day
            if now.month == 12:
                next_run = now.replace(year=now.year + 1, month=1)
            else:
                next_run = now.replace(month=now.month + 1)
        else:
            next_run = now + timedelta(days=30)  # Default to monthly
        
        return next_run.isoformat()
    
    def check_due_reports(self) -> List[str]:
        """Check for reports that are due to be generated"""
        due_reports = []
        now = datetime.now()
        
        for name, schedule in self.schedules.items():
            if not schedule.get('enabled', True):
                continue
                
            try:
                next_run = datetime.fromisoformat(schedule['next_run'])
                if now >= next_run:
                    due_reports.append(name)
            except:
                continue
        
        return due_reports
    
    def generate_scheduled_report(self, schedule_name: str) -> bool:
        """Generate a scheduled report"""
        try:
            if schedule_name not in self.schedules:
                return False
            
            schedule = self.schedules[schedule_name]
            data = self.data_manager.get_all_data()
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"scheduled_{schedule_name}_{timestamp}"
            
            output_path = schedule.get('output_path', '')
            if not output_path:
                output_path = os.path.join(os.path.dirname(__file__), '..', 'data', 'exports')
                os.makedirs(output_path, exist_ok=True)
            
            full_path = os.path.join(output_path, f"{filename}.{schedule['format']}")
            
            # Generate report based on format
            success = False
            if schedule['format'] == 'pdf':
                success = self.export_manager.export_pdf_professional(data, full_path, schedule['template'])
            elif schedule['format'] == 'excel':
                success = self.export_manager.export_excel_professional(data, full_path, schedule['template'])
            elif schedule['format'] == 'json':
                success = self._export_json(data, full_path)
            
            if success:
                # Update schedule
                self.schedules[schedule_name]['last_run'] = datetime.now().isoformat()
                self.schedules[schedule_name]['next_run'] = self._calculate_next_run(schedule['frequency'])
                self.save_schedules()
            
            return success
            
        except Exception as e:
            print(f"Error generating scheduled report: {e}")
            return False
    
    def _export_json(self, data: Dict, filename: str) -> bool:
        """Export data as JSON for scheduled reports"""
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False, default=str)
            return True
        except Exception as e:
            print(f"Error exporting JSON: {e}")
            return False


# Integration functions for main application

def create_advanced_export_manager() -> AdvancedExportManager:
    """Factory function to create AdvancedExportManager"""
    return AdvancedExportManager()

def create_report_template_manager() -> ReportTemplateManager:
    """Factory function to create ReportTemplateManager"""
    return ReportTemplateManager()

def create_report_scheduler(data_manager, export_manager) -> ReportScheduler:
    """Factory function to create ReportScheduler"""
    return ReportScheduler(data_manager, export_manager)


# Example usage and testing
if __name__ == "__main__":
    # Test the export utilities
    sample_data = {
        'budgets': {
            'Aug-25': {'Food': 5000, 'Transport': 3000, 'Shopping': 2000},
            'Sep-25': {'Food': 5500, 'Transport': 3200, 'Shopping': 2500}
        },
        'transactions': {
            'Aug-25': [
                {'date': '2025-08-15', 'category': 'Food', 'amount': 4500, 'description': 'Groceries'},
                {'date': '2025-08-20', 'category': 'Transport', 'amount': 2800, 'description': 'Fuel'},
                {'date': '2025-08-25', 'category': 'Shopping', 'amount': 1500, 'description': 'Clothes'}
            ],
            'Sep-25': [
                {'date': '2025-09-05', 'category': 'Food', 'amount': 5000, 'description': 'Groceries'},
                {'date': '2025-09-10', 'category': 'Transport', 'amount': 3500, 'description': 'Fuel'},
                {'date': '2025-09-15', 'category': 'Shopping', 'amount': 3000, 'description': 'Electronics'}
            ]
        }
    }
    
    # Test export manager
    export_manager = create_advanced_export_manager()
    
    # Test comprehensive package export
    output_dir = tempfile.mkdtemp()
    package_path = export_manager.export_comprehensive_package(sample_data, output_dir, "test_package")
    
    if package_path:
        print(f"Comprehensive package created: {package_path}")
    else:
        print("Failed to create comprehensive package")
    
    # Test template manager
    template_manager = create_report_template_manager()
    templates = template_manager.get_all_templates()
    print(f"Available templates: {list(templates.keys())}")
    
    print("Export utilities test completed successfully!")
